import tkinter as tk
import synonyms as sy
import openpyxl as xl
from tkinter import filedialog
database=[]
answer=[]
result=[]
win1=tk.Tk()
win1.geometry("640x640")
win1.title("校园信息自助查询机器人主界面")
v1=tk.StringVar()
vv1=tk.StringVar()
vv2=tk.StringVar()
def hit1():
    global file
    global database
    global answer
    file=filedialog.askopenfilename()
    v1.set(file)
    count_row=1
    wb=xl.load_workbook(file)
    st1=wb.worksheets[0]
    for row in st1.rows:
            if st1.cell(count_row,1).value !=None:
                database.append(str(st1.cell(count_row, 1).value))
                answer.append(str(st1.cell(count_row, 2).value))
                count_row +=1
    print(database)
    print(answer)
def hit2():
    if file != "":
        result=[]
        global c
        global d
        question=e1.get(1.0,tk.END)
        print(question)
        for i in database:
            r=sy.compare(i,question,seg=False)
            result.append(r)
            print(r)
        print(result)
        win2=tk.Tk()
        win2.geometry("640x480")
        win2.title("问题界面")

        if max(result)==1:
            location=result.index(max(result))
            print(location)
            c=str(database[location])
            c="您的问题是："+c
            print(c)
            d=str(answer[location])
            d="您问题的答案是："+d
            print(d)
            e2=tk.Text(win2,width=60,height=10)
            e2.insert(1.0,c)
            e2.insert(tk.END,"\n")
            e2.insert(tk.END,d)
            e2.pack()
            def hit3():
                win2.destroy()
            b3=tk.Button(win2,text='退出',width=15,height=2,command=hit3)
            b3.pack()
            win2.mainloop()
        if max(result)<0.6:
            location=result.index(max(result))
            print(location)
            e2=tk.Text(win2,width=60,height=10)
            e2.insert(1.0,"抱歉没有找到您要询问的问题，已为您找到校园问题库中最接近的问题，如果是您要问的问题请按确定按键，如果不是请按退出按键后重新提问")
            e2.insert(tk.END,"\n")
            c=str(database[location])
            c="您要问的问题是否是："+c
            print(c)
            e2.insert(tk.END,c)
            e2.pack()
            def hit3():
                win2.destroy()
            def hit4():
                win3=tk.Tk()
                win3.geometry("640x480")
                win3.title("答案界面")
                e3=tk.Text(win3,width=60,height=10)
                s="您问题的答案是：" + answer[location]
                e3.insert(tk.END,s)
                e3.pack()
                win3.mainloop()
            bb4=tk.Button(win2,text='确定',width=15,height=2,command=hit4)
            bb4.pack()
            b3=tk.Button(win2,text='退出',width=15,height=2,command=hit3)
            b3.pack()
            win2.mainloop()
        if max(result)<1 and max(result)>=0.6:
            e2=tk.Text(win2,width=60,height=10)
            e2.insert(1.0,"为您找到3条相似问题，请按按键进行选择，如果没有，请按退出按键后重新提问")
            e2.insert(tk.END,"\n")
            weizhi1=result.index(max(result))
            print(weizhi1)
            c=str(database[weizhi1])
            c="A："+c
            print(c)
            e2.insert(tk.END,c)
            e2.insert(tk.END,"\n")
            result[weizhi1]=0
            weizhi2=result.index(max(result))
            print(weizhi2)
            c=str(database[weizhi2])
            c="B："+c
            print(c)
            e2.insert(tk.END,c)
            e2.insert(tk.END,"\n")
            result[weizhi2]=0
            weizhi3=result.index(max(result))
            print(weizhi3)
            c=str(database[weizhi3])
            c="C："+c
            print(c)
            e2.insert(tk.END,c)
            e2.insert(tk.END,"\n")
            result[weizhi3]=0
            e2.pack()
            def hit3():
                win2.destroy()
            def hitb1():
                win3=tk.Tk()
                win3.geometry("640x480")
                win3.title("答案界面")
                e3=tk.Text(win3,width=60,height=10)
                s="您问题的答案是：" + answer[weizhi1]
                e3.insert(tk.END,s)
                e3.pack()
                win3.mainloop()
            def hitb2():
                win3=tk.Tk()
                win3.geometry("640x480")
                win3.title("答案界面")
                e3=tk.Text(win3,width=60,height=10)
                s="您问题的答案是：" + answer[weizhi2]
                e3.insert(tk.END,s)
                e3.pack()
                win3.mainloop()
            def hitb3():
                win3=tk.Tk()
                win3.geometry("640x480")
                win3.title("答案界面")
                e3=tk.Text(win3,width=60,height=10)
                s="您问题的答案是：" + answer[weizhi3]
                e3.insert(tk.END,s)
                e3.pack()
                win3.mainloop()
            bb1=tk.Button(win2,text='问题A',width=15,height=2,command=hitb1)
            bb1.pack()
            bb2=tk.Button(win2,text='问题B',width=15,height=2,command=hitb2)
            bb2.pack()
            bb3=tk.Button(win2,text='问题C',width=15,height=2,command=hitb3)
            bb3.pack()
            b3=tk.Button(win2,text='退出',width=15,height=2,command=hit3)
            b3.pack()
            win2.mainloop()
    else:
        winwarn=tk.Tk()
        winwarn.geometry("480x240")
        winwarn.title("报错")
        warn=tk.Label(winwarn,text="请载入问题库",font=('Arial',12),width=15,height=2)
        warn.pack()
        warn.mainloop()
def hit4():
    win1.destroy()
b1=tk.Button(win1,text='选择问题库',width=15,height=2,command=hit1)
b1.pack()
l1=tk.Label(win1,textvariable=v1,bg='white',font=('Arial',20),width=30,height=2)
l1.pack()
e1=tk.Text(win1)
e1.pack()
b2=tk.Button(win1,text='提问',width=15,height=2,command=hit2)
b2.pack()
b4=tk.Button(win1,text='退出',width=15,height=2,command=hit4)
b4.pack()
win1.mainloop()
